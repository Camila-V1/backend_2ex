from rest_framework import viewsets, filters, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q, Count
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import AuditLog
from .serializers import AuditLogSerializer, AuditLogFilterSerializer


class AuditLogPagination(PageNumberPagination):
    """Paginación personalizada para logs de auditoría"""
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet para consultar registros de auditoría
    
    **Endpoints:**
    - GET /api/audit/ - Listar registros con filtros
    - GET /api/audit/{id}/ - Detalle de un registro
    - GET /api/audit/stats/ - Estadísticas de auditoría
    - GET /api/audit/export_pdf/ - Exportar a PDF
    - GET /api/audit/export_excel/ - Exportar a Excel
    
    **Filtros disponibles:**
    - action: Tipo de acción
    - severity: Nivel de severidad
    - user_id: ID del usuario
    - username: Nombre de usuario (búsqueda parcial)
    - ip_address: Dirección IP
    - object_type: Tipo de objeto
    - object_id: ID del objeto
    - start_date: Fecha inicial
    - end_date: Fecha final
    - success: true/false
    - search: Búsqueda en descripción, path, error_message
    """
    
    queryset = AuditLog.objects.all()
    serializer_class = AuditLogSerializer
    permission_classes = [permissions.IsAuthenticated, permissions.IsAdminUser]
    pagination_class = AuditLogPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['description', 'path', 'error_message', 'username']
    ordering_fields = ['timestamp', 'action', 'severity', 'username']
    ordering = ['-timestamp']
    
    def get_queryset(self):
        """Aplicar filtros personalizados"""
        queryset = super().get_queryset()
        
        # Filtros desde query params
        action = self.request.query_params.get('action')
        severity = self.request.query_params.get('severity')
        user_id = self.request.query_params.get('user_id')
        username = self.request.query_params.get('username')
        ip_address = self.request.query_params.get('ip_address')
        object_type = self.request.query_params.get('object_type')
        object_id = self.request.query_params.get('object_id')
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        success = self.request.query_params.get('success')
        
        if action:
            queryset = queryset.filter(action=action)
        
        if severity:
            queryset = queryset.filter(severity=severity)
        
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        
        if username:
            queryset = queryset.filter(username__icontains=username)
        
        if ip_address:
            queryset = queryset.filter(ip_address=ip_address)
        
        if object_type:
            queryset = queryset.filter(object_type=object_type)
        
        if object_id:
            queryset = queryset.filter(object_id=object_id)
        
        if start_date:
            queryset = queryset.filter(timestamp__gte=start_date)
        
        if end_date:
            queryset = queryset.filter(timestamp__lte=end_date)
        
        if success is not None:
            success_bool = success.lower() in ['true', '1', 'yes']
            queryset = queryset.filter(success=success_bool)
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """
        Obtener estadísticas de auditoría
        GET /api/audit/stats/
        """
        queryset = self.get_queryset()
        
        # Estadísticas generales
        total_logs = queryset.count()
        
        # Por acción
        by_action = queryset.values('action').annotate(count=Count('id')).order_by('-count')[:10]
        
        # Por severidad
        by_severity = queryset.values('severity').annotate(count=Count('id'))
        
        # Por usuario
        by_user = queryset.values('username').annotate(count=Count('id')).order_by('-count')[:10]
        
        # Por IP
        by_ip = queryset.values('ip_address').annotate(count=Count('id')).order_by('-count')[:10]
        
        # Éxitos vs Errores
        success_count = queryset.filter(success=True).count()
        error_count = queryset.filter(success=False).count()
        
        # Últimas 24 horas
        last_24h = queryset.filter(
            timestamp__gte=timezone.now() - timedelta(hours=24)
        ).count()
        
        # Última semana
        last_week = queryset.filter(
            timestamp__gte=timezone.now() - timedelta(days=7)
        ).count()
        
        return Response({
            'total_logs': total_logs,
            'last_24_hours': last_24h,
            'last_week': last_week,
            'success_count': success_count,
            'error_count': error_count,
            'by_action': by_action,
            'by_severity': by_severity,
            'by_user': by_user,
            'by_ip': by_ip
        })
    
    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        """
        Exportar registros de auditoría a PDF
        GET /api/audit/export_pdf/?[filtros]
        """
        queryset = self.get_queryset()[:1000]  # Limitar a 1000 registros
        
        # Crear PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, 
                                rightMargin=30, leftMargin=30,
                                topMargin=30, bottomMargin=18)
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Contenido
        elements = []
        
        # Título
        title = Paragraph("📋 Reporte de Auditoría del Sistema", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Información del reporte
        info_style = styles['Normal']
        info = Paragraph(
            f"<b>Generado:</b> {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>"
            f"<b>Total de registros:</b> {queryset.count()}<br/>"
            f"<b>Usuario:</b> {request.user.username}",
            info_style
        )
        elements.append(info)
        elements.append(Spacer(1, 20))
        
        # Tabla de datos
        data = [['Fecha/Hora', 'Acción', 'Usuario', 'IP', 'Estado']]
        
        for log in queryset:
            data.append([
                log.timestamp.strftime('%Y-%m-%d %H:%M'),
                log.get_action_display()[:30],
                log.username[:20],
                log.ip_address or 'N/A',
                '✓' if log.success else '✗'
            ])
        
        table = Table(data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 1.3*inch, 0.7*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        
        # Construir PDF
        doc.build(elements)
        
        # Devolver respuesta
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="auditoria_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        return response
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        Exportar registros de auditoría a Excel
        GET /api/audit/export_excel/?[filtros]
        """
        queryset = self.get_queryset()[:5000]  # Limitar a 5000 registros
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Auditoría"
        
        # Estilos
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['ID', 'Fecha/Hora', 'Acción', 'Severidad', 'Usuario', 'IP', 
                   'Método', 'Ruta', 'Descripción', 'Objeto', 'Estado', 'Error']
        ws.append(headers)
        
        # Aplicar estilo a headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos
        for log in queryset:
            ws.append([
                log.id,
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                log.get_action_display(),
                log.get_severity_display(),
                log.username,
                log.ip_address or '',
                log.method,
                log.path,
                log.description[:100],
                f"{log.object_type} #{log.object_id}" if log.object_type else '',
                'Exitoso' if log.success else 'Error',
                log.error_message[:100] if log.error_message else ''
            ])
        
        # Aplicar bordes y ajustar anchos
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # Ajustar anchos de columna
        column_widths = [8, 18, 20, 12, 15, 15, 8, 30, 30, 15, 10, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'
        
        # Guardar en memoria
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Devolver respuesta
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="auditoria_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
