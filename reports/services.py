from io import BytesIO
from datetime import datetime, timedelta
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from decimal import Decimal
import openpyxl
from django.db.models import Count, Sum, Q, F, Avg

from shop_orders.models import Order, OrderItem
from products.models import Product


def generate_sales_report_pdf(start_date, end_date):
    """
    Genera un reporte de ventas en formato PDF.
    Incluye TODO el día de end_date (hasta las 23:59:59).
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    
    # Convertir end_date al día siguiente para incluir todo el día
    end_date_inclusive = end_date + timedelta(days=1)
    
    orders = Order.objects.filter(
        created_at__gte=start_date,
        created_at__lt=end_date_inclusive,
        status=Order.OrderStatus.PAID
    )

    # Datos para la tabla
    data = [['ID Orden', 'Usuario', 'Fecha', 'Total']]
    for order in orders:
        data.append([
            order.id,
            order.user.username,
            order.created_at.strftime('%Y-%m-%d'),
            f"${order.total_price}"
        ])

    table = Table(data)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)

    # Construir el PDF
    elements = [table]
    doc.build(elements)
    
    buffer.seek(0)
    return buffer


def generate_sales_report_excel(start_date, end_date):
    """
    Genera un reporte de ventas en formato Excel.
    Incluye TODO el día de end_date (hasta las 23:59:59).
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Reporte de Ventas"

    headers = ['ID Orden', 'Usuario', 'Email', 'Fecha', 'Estado', 'Total']
    sheet.append(headers)
    
    # Convertir end_date al día siguiente para incluir todo el día
    end_date_inclusive = end_date + timedelta(days=1)
    
    orders = Order.objects.filter(
        created_at__gte=start_date,
        created_at__lt=end_date_inclusive,
        status=Order.OrderStatus.PAID
    )

    for order in orders:
        sheet.append([
            order.id,
            order.user.username,
            order.user.email,
            order.created_at.strftime('%Y-%m-%d %H:%M'),
            order.get_status_display(),
            order.total_price
        ])
    
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def generate_products_report_pdf():
    """
    Genera un reporte de productos en formato PDF.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    
    products = Product.objects.all().select_related('category')

    # Datos para la tabla
    data = [['ID', 'Nombre', 'Categoría', 'Precio', 'Stock', 'Estado']]
    for product in products:
        data.append([
            product.id,
            product.name,
            product.category.name if product.category else 'Sin categoría',
            f"${product.price}",
            product.stock,
            'Activo' if product.is_active else 'Inactivo'
        ])

    table = Table(data)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)

    # Construir el PDF
    elements = [table]
    doc.build(elements)
    
    buffer.seek(0)
    return buffer


def generate_products_report_excel():
    """
    Genera un reporte de productos en formato Excel.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Reporte de Productos"

    headers = ['ID', 'Nombre', 'Categoría', 'Descripción', 'Precio', 'Stock', 'Estado']
    sheet.append(headers)
    
    products = Product.objects.all().select_related('category')

    for product in products:
        sheet.append([
            product.id,
            product.name,
            product.category.name if product.category else 'Sin categoría',
            product.description[:100] if product.description else '',
            product.price,
            product.stock,
            'Activo' if product.is_active else 'Inactivo'
        ])
    
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def generate_invoice_pdf(order):
    """
    Genera un comprobante de venta (factura/nota) en formato PDF para una sola orden.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    
    elements = []

    # Título
    elements.append(Paragraph(f"<b>COMPROBANTE DE VENTA</b>", styles['h1']))
    elements.append(Paragraph(f"Orden #{order.id}", styles['h2']))
    elements.append(Spacer(1, 24))

    # Información del Cliente y Orden
    info_style = styles['Normal']
    elements.append(Paragraph(f"<b>Cliente:</b> {order.user.get_full_name() or order.user.username}", info_style))
    elements.append(Paragraph(f"<b>Email:</b> {order.user.email}", info_style))
    elements.append(Paragraph(f"<b>Fecha de Compra:</b> {order.created_at.strftime('%d/%m/%Y %H:%M:%S')}", info_style))
    elements.append(Paragraph(f"<b>Estado:</b> {order.get_status_display()}", info_style))
    elements.append(Spacer(1, 24))

    # Tabla con los detalles de los productos
    data = [['Producto', 'Cantidad', 'Precio Unitario', 'Subtotal']]
    
    total_general = Decimal('0.00')
    for item in order.items.all():
        subtotal = Decimal(str(item.quantity)) * Decimal(str(item.price))
        total_general += subtotal
        data.append([
            item.product.name,
            str(item.quantity),
            f"${Decimal(str(item.price)):.2f}",
            f"${subtotal:.2f}"
        ])
    
    # Fila del total
    data.append(['', '', '<b>TOTAL</b>', f'<b>${total_general:.2f}</b>'])

    table = Table(data, colWidths=[280, 60, 100, 80])
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        # Estilos para la fila del total
        ('BACKGROUND', (-2, -1), (-1, -1), colors.grey),
        ('TEXTCOLOR', (-2, -1), (-1, -1), colors.whitesmoke),
        ('ALIGN', (-2, -1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (-2, -1), (-1, -1), 'Helvetica-Bold'),
    ])
    table.setStyle(style)
    
    elements.append(table)
    elements.append(Spacer(1, 24))
    
    # Pie de página
    elements.append(Paragraph("<i>Gracias por su compra - SmartSales365</i>", styles['Normal']))
    
    doc.build(elements)
    
    buffer.seek(0)
    return buffer


def generate_dynamic_report_pdf(data, headers, title="Reporte Dinámico"):
    """
    Genera un reporte dinámico en formato PDF con datos y encabezados personalizados.
    
    Args:
        data: Lista de listas con los datos del reporte
        headers: Lista con los nombres de las columnas
        title: Título del reporte
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    
    elements = []
    
    # Título
    elements.append(Paragraph(f"<b>{title}</b>", styles['h1']))
    elements.append(Spacer(1, 24))
    
    # Preparar datos para la tabla
    table_data = [headers]
    table_data.extend(data)
    
    # Crear tabla
    table = Table(table_data)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    return buffer


def generate_dynamic_report_excel(data, headers, title="Reporte Dinámico"):
    """
    Genera un reporte dinámico en formato Excel con datos y encabezados personalizados.
    
    Args:
        data: Lista de listas con los datos del reporte
        headers: Lista con los nombres de las columnas
        title: Título del reporte (usado como nombre de la hoja)
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = title[:31]  # Excel limita a 31 caracteres
    
    # Agregar encabezados
    sheet.append(headers)
    
    # Agregar datos
    for row in data:
        sheet.append(row)
    
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_dynamic_sales_query(parsed_instructions):
    """
    Construye dinámicamente una consulta SQL basada en las instrucciones parseadas.
    
    Args:
        parsed_instructions: Diccionario con las instrucciones parseadas del prompt
            - group_by: 'product', 'customer', None
            - show_customer_names: bool
            - show_product_names: bool
            - count_orders: bool
            - sum_totals: bool
            - start_date: date object
            - end_date: date object
    
    Returns:
        Tupla (queryset, headers, data_formatter)
    """
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    start_date = parsed_instructions.get('start_date')
    end_date = parsed_instructions.get('end_date')
    group_by = parsed_instructions.get('group_by')
    show_customer_names = parsed_instructions.get('show_customer_names', False)
    show_product_names = parsed_instructions.get('show_product_names', False)
    count_orders = parsed_instructions.get('count_orders', False)
    sum_totals = parsed_instructions.get('sum_totals', False)
    
    # Convertir end_date al día siguiente para incluir todo el día
    end_date_inclusive = end_date + timedelta(days=1)
    
    # Caso 1: Agrupado por PRODUCTO
    if group_by == 'product':
        # Consulta agrupada por producto con agregaciones
        queryset = (
            OrderItem.objects
            .filter(
                order__created_at__gte=start_date,
                order__created_at__lt=end_date_inclusive,
                order__status=Order.OrderStatus.PAID
            )
            .values('product__id', 'product__name')
            .annotate(
                total_quantity=Sum('quantity'),
                total_sales=Sum(F('quantity') * F('price')),
                order_count=Count('order__id', distinct=True)
            )
            .order_by('-total_sales')
        )
        
        headers = ['Producto', 'Cantidad Vendida', 'Total Ventas', 'Órdenes']
        
        def formatter(item):
            return [
                item['product__name'] or 'Producto eliminado',
                item['total_quantity'],
                f"${item['total_sales']:.2f}",
                item['order_count']
            ]
        
        return queryset, headers, formatter
    
    # Caso 2: Agrupado por CLIENTE
    elif group_by == 'customer':
        # Consulta agrupada por cliente con agregaciones
        queryset = (
            Order.objects
            .filter(
                created_at__gte=start_date,
                created_at__lt=end_date_inclusive,
                status=Order.OrderStatus.PAID
            )
            .values('user__id', 'user__username', 'user__email', 'user__first_name', 'user__last_name')
            .annotate(
                order_count=Count('id'),
                total_spent=Sum('total_price'),
                avg_order_value=Avg('total_price')
            )
            .order_by('-total_spent')
        )
        
        headers = ['Cliente', 'Email', 'Cantidad de Órdenes', 'Total Gastado', 'Promedio por Orden']
        
        def formatter(item):
            full_name = f"{item['user__first_name']} {item['user__last_name']}".strip()
            display_name = full_name if full_name else item['user__username']
            
            return [
                display_name,
                item['user__email'],
                item['order_count'],
                f"${item['total_spent']:.2f}",
                f"${item['avg_order_value']:.2f}"
            ]
        
        return queryset, headers, formatter
    
    # Caso 3: Reporte DETALLADO (con nombres de clientes y/o productos)
    else:
        # Consulta detallada de todas las órdenes
        queryset = (
            Order.objects
            .filter(
                created_at__gte=start_date,
                created_at__lt=end_date_inclusive,
                status=Order.OrderStatus.PAID
            )
            .select_related('user')
            .prefetch_related('items__product')
            .order_by('-created_at')
        )
        
        # Construir encabezados dinámicamente
        headers = ['ID Orden']
        
        if show_customer_names:
            headers.extend(['Cliente', 'Email'])
        
        headers.append('Fecha')
        
        if show_product_names:
            headers.append('Productos')
        
        headers.append('Total')
        
        def formatter(order):
            row = [order.id]
            
            if show_customer_names:
                full_name = order.user.get_full_name() or order.user.username
                row.extend([full_name, order.user.email])
            
            row.append(order.created_at.strftime('%Y-%m-%d %H:%M'))
            
            if show_product_names:
                products_list = ", ".join([
                    f"{item.product.name} (x{item.quantity})" 
                    for item in order.items.all()
                    if item.product
                ])
                row.append(products_list or 'N/A')
            
            row.append(f"${order.total_price:.2f}")
            
            return row
        
        return queryset, headers, formatter
